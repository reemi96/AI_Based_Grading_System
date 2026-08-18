import json
import os
import re
import threading
import urllib.error
import urllib.request
from pathlib import Path

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, field_validator


app = FastAPI(title="AI Based Grading System Backend")
MOCK_MODE = True
generation_lock = threading.Lock()
excel_lock = threading.Lock()
MASTER_DATASET_ROOT = Path(__file__).resolve().parent.parent / "Master_Project_Dataset"
COLAB_GRADE_URL = os.getenv(
    "COLAB_GRADE_URL",
    "https://love-single-purpose-insights.trycloudflare.com/grade",
)
COLAB_TIMEOUT_SECONDS = float(os.getenv("COLAB_TIMEOUT_SECONDS", "60"))
SUBJECT_FOLDERS = {
    "Programming Languages": "Programming Languages_Dataset",
    "Programming 1": "Programming1_Dataset",
    "Web Programming": "Web Programming_Dataset",
}
RECORDS_FILE = Path(__file__).with_name("grading_records.xlsx")
RECORD_COLUMNS = [
    "Saved At",
    "Subject",
    "Question ID",
    "Rubric File",
    "Student File",
    "AI Grade",
    "Final Grade",
    "Instructor Notes",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:5173",
        "http://localhost:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class GradeRequest(BaseModel):
    rubric: str = Field(..., min_length=1)
    student_answer: str = Field(..., min_length=1)


class GradeResponse(BaseModel):
    predicted_grade: str
    raw_response: str


class GradeAssignmentRequest(BaseModel):
    subject: str
    question_id: str
    student_answer: str
    rubric_text: str
    rubric_file_name: str = ""
    student_file_name: str = ""

    @field_validator("subject")
    @classmethod
    def subject_must_be_supported(cls, value):
        cleaned_value = value.strip()
        if cleaned_value not in SUBJECT_FOLDERS:
            raise ValueError("Subject is required and must be one of the supported model subjects.")
        return cleaned_value

    @field_validator("question_id")
    @classmethod
    def question_id_must_be_present(cls, value):
        cleaned_value = value.strip().upper()
        if not re.fullmatch(r"Q\d+", cleaned_value):
            raise ValueError("Question ID is required and must use a format like Q001.")
        return cleaned_value

    @field_validator("student_answer")
    @classmethod
    def student_answer_must_not_be_empty(cls, value):
        if not value.strip():
            raise ValueError("Student answer is empty.")
        return value

    @field_validator("rubric_text")
    @classmethod
    def rubric_text_must_not_be_empty(cls, value):
        if not value.strip():
            raise ValueError("Rubric text is required.")
        return value


class GradeAssignmentResponse(BaseModel):
    predictedGrade: str
    rawResponse: str
    questionId: str


class SaveDecisionRequest(BaseModel):
    subject: str = ""
    question_id: str = ""
    rubric_file_name: str
    student_file_name: str
    ai_grade: str
    final_grade: str = Field(..., min_length=1)
    instructor_notes: str = ""
    saved_at: str

    @field_validator("final_grade")
    @classmethod
    def final_grade_must_be_between_zero_and_ten(cls, value):
        match = re.search(r"\d+(?:\.\d+)?", value.strip())
        if not match:
            raise ValueError("Final grade is required and must be between 0 and 10.")

        grade_value = float(match.group(0))
        if grade_value < 0 or grade_value > 10:
            raise ValueError("Final grade must be between 0 and 10.")

        return value


class SaveDecisionResponse(BaseModel):
    status: str
    message: str


def ensure_record_columns(sheet):
    headers = [cell.value for cell in sheet[1]]
    if not any(headers):
        for column_index, column_name in enumerate(RECORD_COLUMNS, start=1):
            sheet.cell(row=1, column=column_index, value=column_name)
        return RECORD_COLUMNS

    extra_columns = [header for header in headers if header and header not in RECORD_COLUMNS]
    ordered_headers = RECORD_COLUMNS + extra_columns

    if headers == ordered_headers:
        return headers

    rows = list(sheet.iter_rows(values_only=True))
    old_headers = list(rows[0])
    max_columns = max(len(old_headers), len(ordered_headers))

    for column_index, column_name in enumerate(ordered_headers, start=1):
        sheet.cell(row=1, column=column_index, value=column_name)

    for row_index, row_values in enumerate(rows[1:], start=2):
        row_by_header = {
            header: row_values[column_index]
            for column_index, header in enumerate(old_headers)
            if header and column_index < len(row_values)
        }
        for column_index, column_name in enumerate(ordered_headers, start=1):
            sheet.cell(row=row_index, column=column_index, value=row_by_header.get(column_name, ""))

    for row_index in range(1, sheet.max_row + 1):
        for column_index in range(len(ordered_headers) + 1, max_columns + 1):
            sheet.cell(row=row_index, column=column_index, value=None)

    return ordered_headers


def read_text_with_fallbacks(path: Path) -> str:
    encodings = ["utf-8", "utf-8-sig", "cp1256", "cp1252"]
    for encoding in encodings:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue

    return path.read_text(encoding="utf-8", errors="replace")


def find_single_question_file(directory: Path, question_id: str, label: str) -> Path:
    if not directory.exists():
        raise HTTPException(status_code=404, detail=f"{label} folder is missing: {directory}")

    matches = [
        path
        for path in directory.iterdir()
        if path.is_file() and path.name.upper().startswith(question_id.upper())
    ]

    if not matches:
        raise HTTPException(status_code=404, detail=f"{label} file is missing for {question_id}.")

    if len(matches) > 1:
        file_names = ", ".join(path.name for path in matches)
        raise HTTPException(
            status_code=409,
            detail=f"More than one {label.lower()} file starts with {question_id}: {file_names}",
        )

    return matches[0]


def call_colab_grade(payload: dict) -> dict:
    request = urllib.request.Request(
        COLAB_GRADE_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=COLAB_TIMEOUT_SECONDS) as response:
            response_text = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise HTTPException(
            status_code=502,
            detail=f"Colab grading API returned HTTP {exc.code}: {detail}",
        ) from exc
    except urllib.error.URLError as exc:
        raise HTTPException(
            status_code=503,
            detail="Could not connect to Colab grading API. Check that Colab is running and the ngrok URL is current.",
        ) from exc
    except TimeoutError as exc:
        raise HTTPException(status_code=504, detail="Colab grading API request timed out.") from exc

    try:
        return json.loads(response_text)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="Colab grading API returned invalid JSON.") from exc


def build_prompt(rubric: str, student_answer: str) -> str:
    return (
        "\u0635\u062d\u062d \u062d\u0644 \u0627\u0644\u0637\u0627\u0644\u0628 "
        "\u0627\u0644\u0628\u0631\u0645\u062c\u064a \u0627\u0639\u062a\u0645\u0627\u062f\u0627\u064b "
        "\u0639\u0644\u0649 \u0633\u0644\u0645 \u0627\u0644\u062a\u0635\u062d\u064a\u062d "
        "\u0648\u0627\u0644\u062d\u0644 \u0627\u0644\u0646\u0645\u0648\u0630\u062c\u064a.\n\n"
        "\u0635\u062d\u062d \u062d\u0644 \u0627\u0644\u0637\u0627\u0644\u0628 "
        "\u0627\u0639\u062a\u0645\u0627\u062f\u0627\u064b \u0639\u0644\u0649 "
        "\u0627\u0644\u0633\u0624\u0627\u0644\u060c \u0633\u0644\u0645 "
        "\u0627\u0644\u062a\u0635\u062d\u064a\u062d\u060c \u0648\u0627\u0644\u062d\u0644 "
        "\u0627\u0644\u0646\u0645\u0648\u0630\u062c\u064a.\n\n"
        "\u0633\u0644\u0645 \u0627\u0644\u062a\u0635\u062d\u064a\u062d:\n"
        f"{rubric.strip()}\n\n"
        "\u062d\u0644 \u0627\u0644\u0637\u0627\u0644\u0628:\n"
        f"{student_answer.strip()}\n\n"
        "\u0623\u0639\u0637\u0650 \u0627\u0644\u0639\u0644\u0627\u0645\u0629 "
        "\u0627\u0644\u0646\u0647\u0627\u0626\u064a\u0629 \u0641\u0642\u0637 "
        "\u0645\u0646 10."
    )


def get_model_device(model):
    return next(model.parameters()).device


def extract_grade(raw_response: str) -> str:
    grade_match = re.search(r"(\d+(?:\.\d+)?\s*/\s*10)", raw_response)
    if grade_match:
        return grade_match.group(1).replace(" ", "")

    labeled_match = re.search(
        r"(?:\u0627\u0644\u0639\u0644\u0627\u0645\u0629|grade|score)\s*[:\uff1a]?\s*(\d+(?:\.\d+)?)",
        raw_response,
        flags=re.IGNORECASE,
    )
    if labeled_match:
        return f"{labeled_match.group(1)}/10"

    return raw_response.strip()


@app.get("/health")
def health():
    return {"status": "ok", "mock_mode": MOCK_MODE}


@app.post("/grade-assignment", response_model=GradeAssignmentResponse)
def grade_assignment(payload: GradeAssignmentRequest):
    print("Resolved dataset root:", MASTER_DATASET_ROOT)
    print("Subject:", payload.subject)
    print("Question ID:", payload.question_id)

    if not MASTER_DATASET_ROOT.exists():
        raise HTTPException(
            status_code=500,
            detail=f"MASTER_DATASET_ROOT does not exist: {MASTER_DATASET_ROOT}",
        )

    subject_folder = SUBJECT_FOLDERS[payload.subject]
    subject_directory = MASTER_DATASET_ROOT / subject_folder
    if not subject_directory.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Subject folder does not exist: {subject_directory}",
        )

    question_file = find_single_question_file(
        subject_directory / "Questions",
        payload.question_id,
        "Question",
    )
    reference_solution_file = find_single_question_file(
        subject_directory / "Reference Solutions",
        payload.question_id,
        "Reference solution",
    )

    question_text = read_text_with_fallbacks(question_file)
    reference_solution_text = read_text_with_fallbacks(reference_solution_file)

    print("Located question filename:", question_file.name)
    print("Located reference solution filename:", reference_solution_file.name)
    print("Question text length:", len(question_text))
    print("Rubric text length:", len(payload.rubric_text))
    print("Reference solution length:", len(reference_solution_text))
    print("Student answer length:", len(payload.student_answer))

    colab_response = call_colab_grade(
        {
            "subject": payload.subject,
            "qid": payload.question_id,
            "question_text": question_text,
            "rubric": payload.rubric_text,
            "reference_solution": reference_solution_text,
            "student_answer": payload.student_answer,
        }
    )

    predicted_grade = colab_response.get("predictedGrade") or colab_response.get("predicted_grade")
    raw_response = colab_response.get("rawResponse") or colab_response.get("raw_response") or ""

    if predicted_grade is None:
        raise HTTPException(status_code=502, detail="Colab grading API response did not include a predicted grade.")

    return GradeAssignmentResponse(
        predictedGrade=predicted_grade,
        rawResponse=raw_response,
        questionId=payload.question_id,
    )


@app.post("/save-decision", response_model=SaveDecisionResponse)
def save_decision(payload: SaveDecisionRequest):
    try:
        with excel_lock:
            if RECORDS_FILE.exists():
                workbook = load_workbook(RECORDS_FILE)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Grading Records"
                sheet.append(RECORD_COLUMNS)

            headers = ensure_record_columns(sheet)
            record = {
                "Saved At": payload.saved_at,
                "Subject": payload.subject,
                "Question ID": payload.question_id,
                "Rubric File": payload.rubric_file_name,
                "Student File": payload.student_file_name,
                "AI Grade": payload.ai_grade,
                "Final Grade": payload.final_grade,
                "Instructor Notes": payload.instructor_notes,
            }

            sheet.append([record.get(header, "") for header in headers])
            workbook.save(RECORDS_FILE)

        return SaveDecisionResponse(
            status="success",
            message="Final decision saved successfully.",
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Could not save final decision.") from exc


@app.post("/grade", response_model=GradeResponse)
def grade_solution(payload: GradeRequest):
    if MOCK_MODE:
        return GradeResponse(
            predicted_grade="8/10",
            raw_response="Mock response: AI model will be connected later through Colab.",
        )

    try:
        import torch

        try:
            from .model_loader import load_model
        except ImportError:
            from model_loader import load_model

        tokenizer, model = load_model()
        prompt = build_prompt(payload.rubric, payload.student_answer)
        messages = [{"role": "user", "content": prompt}]

        formatted_prompt = tokenizer.apply_chat_template(
            messages,
            tokenize=False,
            add_generation_prompt=True,
        )
        inputs = tokenizer(formatted_prompt, return_tensors="pt").to(get_model_device(model))

        with generation_lock, torch.no_grad():
            generated_ids = model.generate(
                **inputs,
                max_new_tokens=32,
                do_sample=False,
                pad_token_id=tokenizer.eos_token_id,
                eos_token_id=tokenizer.eos_token_id,
            )

        new_tokens = generated_ids[:, inputs["input_ids"].shape[-1] :]
        raw_response = tokenizer.decode(new_tokens[0], skip_special_tokens=True).strip()

        return GradeResponse(
            predicted_grade=extract_grade(raw_response),
            raw_response=raw_response,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
